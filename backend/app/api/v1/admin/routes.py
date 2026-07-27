import hashlib
import io
from pathlib import Path
from typing import Annotated
from uuid import UUID, uuid4
from zoneinfo import ZoneInfo

import anyio
from fastapi import APIRouter, Depends, File, Form, Request, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from PIL import Image, UnidentifiedImageError
from sqlalchemy import func, select, update

from app.core.time import utc_now
from app.exceptions import AppError
from app.models import (
    AttendanceEvent,
    Event,
    ImportBatch,
    ImportRow,
    ImportRowError,
    MapAsset,
    QrToken,
    Registration,
    ScanAttempt,
    ScheduleItem,
    SourceStudent,
    User,
)
from app.models.enums import (
    AttendanceType,
    EventStatus,
    ImportStatus,
    PresenceStatus,
    QrStatus,
)
from app.schemas.admin import (
    ArchiveEventRequest,
    CancelRequest,
    ImportConfirmRequest,
    ScheduleCreateRequest,
    SchedulePatchRequest,
    StudentPatchRequest,
    UserCreateRequest,
    UserPatchRequest,
)
from app.security.crypto import hash_password
from app.security.dependencies import SessionDep, require_admin
from app.services.audit import add_audit
from app.services.imports import apply_import, parse_workbook, stage_import
from app.settings import settings

router = APIRouter(prefix="/admin", tags=["admin"])


def _safe_excel(value: object) -> object:
    if isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _batch_dict(batch: ImportBatch) -> dict[str, object]:
    return {
        "id": batch.id,
        "event_id": batch.event_id,
        "filename": batch.filename,
        "status": batch.status,
        "total_rows": batch.total_rows,
        "valid_rows": batch.valid_rows,
        "error_rows": batch.error_rows,
        "warning_rows": batch.warning_rows,
        "created_rows": batch.created_rows,
        "updated_rows": batch.updated_rows,
        "unchanged_rows": batch.unchanged_rows,
        "deactivated_rows": batch.deactivated_rows,
        "preview_version": batch.preview_version,
        "uploaded_at": batch.uploaded_at,
        "validated_at": batch.validated_at,
    }


@router.post("/imports", status_code=201)
async def upload_import(
    request: Request,
    session: SessionDep,
    user: Annotated[User, Depends(require_admin)],
    event_id: Annotated[UUID, Form()],
    file: Annotated[UploadFile, File()],
) -> dict[str, object]:
    if not file.filename or Path(file.filename).suffix.casefold() != ".xlsx":
        raise AppError("IMPORT_INVALID_EXTENSION", "Разрешены только файлы .xlsx", 415)
    data = await file.read(settings.max_upload_bytes + 1)
    if len(data) > settings.max_upload_bytes:
        raise AppError("IMPORT_FILE_TOO_LARGE", "Файл превышает 20 МиБ", 413)
    event = await session.get(Event, event_id)
    if not event:
        raise AppError("EVENT_NOT_FOUND", "Мероприятие не найдено", 404)
    now = utc_now()
    batch = ImportBatch(
        event_id=event_id,
        filename=Path(file.filename).name[:255],
        file_size=len(data),
        file_hash=hashlib.sha256(data).hexdigest(),
        uploaded_by_user_id=user.id,
        status=ImportStatus.VALIDATING,
        uploaded_at=now,
    )
    session.add(batch)
    await session.commit()
    try:
        parsed, issues = await anyio.to_thread.run_sync(
            parse_workbook, data, settings.max_import_rows
        )
        await stage_import(session, batch, parsed, issues)
        add_audit(
            session,
            request,
            "IMPORT_UPLOAD",
            "ImportBatch",
            str(batch.id),
            user.id,
            {"filename": batch.filename, "rows": batch.total_rows},
        )
        await session.commit()
    except Exception:
        await session.rollback()
        failed_batch = await session.get(ImportBatch, batch.id)
        if failed_batch:
            failed_batch.status = ImportStatus.FAILED
            failed_batch.failed_at = utc_now()
            await session.commit()
        raise
    return _batch_dict(batch)


@router.get("/imports/{import_id}")
async def get_import(
    import_id: UUID, session: SessionDep, _: Annotated[User, Depends(require_admin)]
) -> dict[str, object]:
    batch = await session.get(ImportBatch, import_id)
    if not batch:
        raise AppError("IMPORT_NOT_FOUND", "Импорт не найден", 404)
    return _batch_dict(batch)


@router.get("/imports/{import_id}/preview")
async def import_preview(
    import_id: UUID,
    session: SessionDep,
    _: Annotated[User, Depends(require_admin)],
    limit: int = 100,
    offset: int = 0,
) -> dict[str, object]:
    batch = await session.get(ImportBatch, import_id)
    if not batch:
        raise AppError("IMPORT_NOT_FOUND", "Импорт не найден", 404)
    rows = (
        await session.scalars(
            select(ImportRow)
            .where(ImportRow.import_batch_id == import_id)
            .order_by(ImportRow.row_number)
            .offset(offset)
            .limit(min(limit, 500))
        )
    ).all()
    return {
        "batch": _batch_dict(batch),
        "items": [
            {
                "id": row.id,
                "row_number": row.row_number,
                "source_id": row.source_id,
                "name": row.name,
                "group": row.group,
                "institute": row.institute,
                "action": row.action,
                "validation_status": row.validation_status,
            }
            for row in rows
        ],
    }


@router.post("/imports/{import_id}/confirm")
async def confirm_import(
    import_id: UUID,
    payload: ImportConfirmRequest,
    request: Request,
    session: SessionDep,
    user: Annotated[User, Depends(require_admin)],
) -> dict[str, object]:
    batch = await session.scalar(
        select(ImportBatch).where(ImportBatch.id == import_id).with_for_update()
    )
    if not batch:
        raise AppError("IMPORT_NOT_FOUND", "Импорт не найден", 404)
    if batch.status == ImportStatus.APPLIED:
        return _batch_dict(batch)
    if (
        batch.status != ImportStatus.READY_TO_CONFIRM
        or batch.preview_version != payload.preview_version
    ):
        raise AppError(
            "IMPORT_PREVIEW_STALE", "Предпросмотр устарел или импорт нельзя подтвердить", 409
        )
    if batch.warning_rows and not payload.accept_warnings:
        raise AppError("IMPORT_WARNINGS_NOT_ACCEPTED", "Подтвердите предупреждения", 422)
    if batch.deactivated_rows and not payload.confirm_deactivations:
        raise AppError(
            "IMPORT_DEACTIVATIONS_NOT_CONFIRMED",
            "Подтвердите деактивацию отсутствующих студентов",
            422,
        )
    if batch.deactivated_rows >= 100 and payload.confirmation_phrase != "ДЕАКТИВИРОВАТЬ":
        raise AppError(
            "CONFIRMATION_PHRASE_INVALID", "Введите контрольную фразу ДЕАКТИВИРОВАТЬ", 422
        )
    try:
        batch.status, batch.confirmed_at = ImportStatus.CONFIRMED, utc_now()
        await apply_import(session, batch, payload.confirm_deactivations)
        batch.status = ImportStatus.APPLIED
        add_audit(
            session,
            request,
            "IMPORT_APPLY",
            "ImportBatch",
            str(batch.id),
            user.id,
            {
                "created": batch.created_rows,
                "updated": batch.updated_rows,
                "deactivated": batch.deactivated_rows if payload.confirm_deactivations else 0,
            },
        )
        await session.commit()
    except Exception:
        await session.rollback()
        failed_batch = await session.get(ImportBatch, import_id)
        if failed_batch:
            failed_batch.status = ImportStatus.FAILED
            failed_batch.failed_at = utc_now()
            await session.commit()
        raise
    return _batch_dict(batch)


@router.post("/imports/{import_id}/cancel")
async def cancel_import(
    import_id: UUID,
    payload: CancelRequest,
    request: Request,
    session: SessionDep,
    user: Annotated[User, Depends(require_admin)],
) -> dict[str, object]:
    batch = await session.scalar(
        select(ImportBatch).where(ImportBatch.id == import_id).with_for_update()
    )
    if not batch:
        raise AppError("IMPORT_NOT_FOUND", "Импорт не найден", 404)
    if batch.status == ImportStatus.APPLIED:
        raise AppError("IMPORT_ALREADY_APPLIED", "Применённый импорт нельзя отменить", 409)
    batch.status = ImportStatus.CANCELLED
    add_audit(
        session,
        request,
        "IMPORT_CANCEL",
        "ImportBatch",
        str(batch.id),
        user.id,
        {"reason": payload.reason},
    )
    await session.commit()
    return _batch_dict(batch)


@router.get("/imports/{import_id}/errors.xlsx")
async def import_errors(
    import_id: UUID, session: SessionDep, _: Annotated[User, Depends(require_admin)]
) -> StreamingResponse:
    errors = (
        await session.execute(
            select(ImportRowError, ImportRow)
            .outerjoin(ImportRow, ImportRow.id == ImportRowError.import_row_id)
            .where(ImportRowError.import_batch_id == import_id)
            .order_by(ImportRowError.row_number)
        )
    ).all()
    workbook, sheet = Workbook(), None
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Ошибки"
    sheet.append(["row_number", "id", "name", "group", "institute", "error_code", "error_message"])
    for error, row in errors:
        sheet.append(
            [
                error.row_number,
                _safe_excel(row.source_id if row else ""),
                _safe_excel(row.name if row else ""),
                _safe_excel(row.group if row else ""),
                _safe_excel(row.institute if row else ""),
                error.error_code,
                error.error_message,
            ]
        )
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="import-{import_id}-errors.xlsx"'},
    )


@router.get("/students")
async def students(
    session: SessionDep,
    _: Annotated[User, Depends(require_admin)],
    event_id: UUID,
    search: str | None = None,
    limit: int = 50,
    offset: int = 0,
) -> dict[str, object]:
    query = select(SourceStudent).where(SourceStudent.event_id == event_id)
    if search:
        query = query.where(SourceStudent.normalized_full_name.contains(search.casefold()))
    rows = (
        await session.scalars(
            query.order_by(SourceStudent.full_name).offset(offset).limit(min(limit, 200))
        )
    ).all()
    return {
        "items": [
            {
                "id": row.id,
                "source_id": row.source_id,
                "full_name": row.full_name,
                "study_group": row.study_group,
                "institute": row.institute,
                "is_active": row.is_active,
            }
            for row in rows
        ]
    }


@router.get("/students/{student_id}")
async def student_detail(
    student_id: UUID, session: SessionDep, _: Annotated[User, Depends(require_admin)]
) -> dict[str, object]:
    student = await session.get(SourceStudent, student_id)
    if not student:
        raise AppError("STUDENT_NOT_FOUND", "Студент не найден", 404)
    registration = await session.scalar(
        select(Registration).where(Registration.source_student_id == student.id)
    )
    return {
        "id": student.id,
        "source_id": student.source_id,
        "full_name": student.full_name,
        "study_group": student.study_group,
        "institute": student.institute,
        "is_active": student.is_active,
        "registration": {
            "public_id": registration.public_id,
            "presence_status": registration.presence_status,
        }
        if registration
        else None,
    }


@router.patch("/students/{student_id}")
async def patch_student(
    student_id: UUID,
    payload: StudentPatchRequest,
    request: Request,
    session: SessionDep,
    user: Annotated[User, Depends(require_admin)],
) -> dict[str, object]:
    student = await session.scalar(
        select(SourceStudent).where(SourceStudent.id == student_id).with_for_update()
    )
    if not student:
        raise AppError("STUDENT_NOT_FOUND", "Студент не найден", 404)
    if payload.is_active is not None:
        student.is_active = payload.is_active
        student.deactivated_at = None if payload.is_active else utc_now()
        if not payload.is_active:
            registration_ids = select(Registration.id).where(
                Registration.source_student_id == student.id
            )
            await session.execute(
                update(QrToken)
                .where(
                    QrToken.registration_id.in_(registration_ids), QrToken.status == QrStatus.ACTIVE
                )
                .values(status=QrStatus.REVOKED, revoked_at=utc_now())
            )
    add_audit(
        session,
        request,
        "STUDENT_UPDATE",
        "SourceStudent",
        str(student.id),
        user.id,
        payload.model_dump(exclude_none=True),
    )
    await session.commit()
    return {"id": student.id, "is_active": student.is_active}


@router.get("/users")
async def users(
    session: SessionDep, _: Annotated[User, Depends(require_admin)]
) -> dict[str, object]:
    rows = (await session.scalars(select(User).order_by(User.username))).all()
    return {
        "items": [
            {
                "id": item.id,
                "username": item.username,
                "role": item.role,
                "is_active": item.is_active,
                "last_login_at": item.last_login_at,
            }
            for item in rows
        ]
    }


@router.post("/users", status_code=201)
async def create_user(
    payload: UserCreateRequest,
    request: Request,
    session: SessionDep,
    admin: Annotated[User, Depends(require_admin)],
) -> dict[str, object]:
    if await session.scalar(select(User.id).where(User.username == payload.username)):
        raise AppError("USERNAME_EXISTS", "Пользователь уже существует", 409)
    user = User(
        username=payload.username,
        password_hash=hash_password(payload.password),
        role=payload.role,
        is_active=True,
    )
    session.add(user)
    await session.flush()
    add_audit(
        session,
        request,
        "USER_CREATE",
        "User",
        str(user.id),
        admin.id,
        {"username": user.username, "role": user.role.value},
    )
    await session.commit()
    return {
        "id": user.id,
        "username": user.username,
        "role": user.role,
        "is_active": user.is_active,
    }


@router.patch("/users/{user_id}")
async def patch_user(
    user_id: UUID,
    payload: UserPatchRequest,
    request: Request,
    session: SessionDep,
    admin: Annotated[User, Depends(require_admin)],
) -> dict[str, object]:
    target = await session.scalar(select(User).where(User.id == user_id).with_for_update())
    if not target:
        raise AppError("USER_NOT_FOUND", "Пользователь не найден", 404)
    if target.id == admin.id and payload.is_active is False:
        raise AppError(
            "SELF_DEACTIVATION_FORBIDDEN", "Нельзя деактивировать собственную учётную запись", 409
        )
    if payload.is_active is not None:
        target.is_active = payload.is_active
    if payload.password:
        target.password_hash = hash_password(payload.password)
    add_audit(
        session,
        request,
        "USER_UPDATE",
        "User",
        str(target.id),
        admin.id,
        {"is_active": payload.is_active, "password_changed": bool(payload.password)},
    )
    await session.commit()
    return {
        "id": target.id,
        "username": target.username,
        "role": target.role,
        "is_active": target.is_active,
    }


@router.get("/schedule")
async def admin_schedule(
    event_id: UUID, session: SessionDep, _: Annotated[User, Depends(require_admin)]
) -> dict[str, object]:
    rows = (
        await session.scalars(
            select(ScheduleItem)
            .where(ScheduleItem.event_id == event_id)
            .order_by(ScheduleItem.starts_at)
        )
    ).all()
    return {
        "items": [
            {
                "id": row.id,
                "title": row.title,
                "description": row.description,
                "location": row.location,
                "starts_at": row.starts_at,
                "ends_at": row.ends_at,
                "display_order": row.display_order,
                "is_published": row.is_published,
            }
            for row in rows
        ]
    }


@router.post("/schedule", status_code=201)
async def create_schedule(
    payload: ScheduleCreateRequest,
    request: Request,
    session: SessionDep,
    user: Annotated[User, Depends(require_admin)],
) -> dict[str, object]:
    if payload.ends_at <= payload.starts_at:
        raise AppError("SCHEDULE_TIME_INVALID", "Окончание должно быть позже начала", 422)
    item = ScheduleItem(**payload.model_dump())
    session.add(item)
    await session.flush()
    add_audit(session, request, "SCHEDULE_CREATE", "ScheduleItem", str(item.id), user.id)
    await session.commit()
    return {"id": item.id}


@router.patch("/schedule/{item_id}")
async def patch_schedule(
    item_id: UUID,
    payload: SchedulePatchRequest,
    request: Request,
    session: SessionDep,
    user: Annotated[User, Depends(require_admin)],
) -> dict[str, object]:
    item = await session.get(ScheduleItem, item_id)
    if not item:
        raise AppError("SCHEDULE_NOT_FOUND", "Элемент расписания не найден", 404)
    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(item, field, value)
    if item.ends_at <= item.starts_at:
        raise AppError("SCHEDULE_TIME_INVALID", "Окончание должно быть позже начала", 422)
    add_audit(
        session,
        request,
        "SCHEDULE_UPDATE",
        "ScheduleItem",
        str(item.id),
        user.id,
        payload.model_dump(exclude_unset=True, mode="json"),
    )
    await session.commit()
    return {"id": item.id}


@router.delete("/schedule/{item_id}", status_code=204)
async def delete_schedule(
    item_id: UUID,
    request: Request,
    session: SessionDep,
    user: Annotated[User, Depends(require_admin)],
) -> None:
    item = await session.get(ScheduleItem, item_id)
    if not item:
        raise AppError("SCHEDULE_NOT_FOUND", "Элемент расписания не найден", 404)
    add_audit(session, request, "SCHEDULE_DELETE", "ScheduleItem", str(item.id), user.id)
    await session.delete(item)
    await session.commit()


@router.post("/map", status_code=201)
async def upload_map(
    request: Request,
    session: SessionDep,
    user: Annotated[User, Depends(require_admin)],
    event_id: Annotated[UUID, Form()],
    file: Annotated[UploadFile, File()],
) -> dict[str, object]:
    data = await file.read(10 * 1024 * 1024 + 1)
    if len(data) > 10 * 1024 * 1024:
        raise AppError("MAP_FILE_TOO_LARGE", "Карта превышает 10 МиБ", 413)
    try:
        image = Image.open(io.BytesIO(data))
        image.load()
    except (UnidentifiedImageError, OSError) as exc:
        raise AppError("MAP_INVALID", "Файл не является безопасным изображением", 422) from exc
    if image.width * image.height > 100_000_000 or image.width > 20_000 or image.height > 20_000:
        raise AppError("MAP_DIMENSIONS_TOO_LARGE", "Слишком большое разрешение карты", 422)
    directory = Path(settings.upload_dir) / "maps"
    directory.mkdir(parents=True, exist_ok=True)
    path = directory / f"{uuid4()}.webp"
    image.convert("RGB").save(path, format="WEBP", quality=92)
    encoded = path.read_bytes()
    await session.execute(
        update(MapAsset)
        .where(MapAsset.event_id == event_id, MapAsset.is_active.is_(True))
        .values(is_active=False)
    )
    asset = MapAsset(
        event_id=event_id,
        storage_path=str(path),
        mime_type="image/webp",
        file_size=len(encoded),
        original_name=Path(file.filename or "map").name[:255],
        file_hash=hashlib.sha256(encoded).hexdigest(),
        is_active=True,
        uploaded_at=utc_now(),
        uploaded_by_user_id=user.id,
        width=image.width,
        height=image.height,
    )
    session.add(asset)
    await session.flush()
    add_audit(session, request, "MAP_UPLOAD", "MapAsset", str(asset.id), user.id)
    await session.commit()
    return {"id": asset.id, "width": asset.width, "height": asset.height}


@router.get("/statistics/summary")
async def statistics_summary(
    event_id: UUID, session: SessionDep, _: Annotated[User, Depends(require_admin)]
) -> dict[str, object]:
    total = (
        await session.scalar(
            select(func.count())
            .select_from(SourceStudent)
            .where(SourceStudent.event_id == event_id)
        )
        or 0
    )
    active = (
        await session.scalar(
            select(func.count())
            .select_from(SourceStudent)
            .where(SourceStudent.event_id == event_id, SourceStudent.is_active.is_(True))
        )
        or 0
    )
    registered = (
        await session.scalar(
            select(func.count())
            .select_from(Registration)
            .where(Registration.event_id == event_id, Registration.revoked_at.is_(None))
        )
        or 0
    )
    inside = (
        await session.scalar(
            select(func.count())
            .select_from(Registration)
            .where(
                Registration.event_id == event_id,
                Registration.presence_status == PresenceStatus.INSIDE,
            )
        )
        or 0
    )
    unique_visitors = (
        await session.scalar(
            select(func.count(func.distinct(AttendanceEvent.registration_id))).where(
                AttendanceEvent.event_id == event_id,
                AttendanceEvent.event_type == AttendanceType.ENTRY,
            )
        )
        or 0
    )
    entries = (
        await session.scalar(
            select(func.count())
            .select_from(AttendanceEvent)
            .where(
                AttendanceEvent.event_id == event_id,
                AttendanceEvent.event_type == AttendanceType.ENTRY,
            )
        )
        or 0
    )
    exits = (
        await session.scalar(
            select(func.count())
            .select_from(AttendanceEvent)
            .where(
                AttendanceEvent.event_id == event_id,
                AttendanceEvent.event_type == AttendanceType.EXIT,
            )
        )
        or 0
    )
    errors = (
        await session.scalar(
            select(func.count())
            .select_from(ScanAttempt)
            .where(ScanAttempt.event_id == event_id, ScanAttempt.result == "ERROR")
        )
        or 0
    )
    return {
        "as_of": utc_now(),
        "total_students": total,
        "active_students": active,
        "registered": registered,
        "unique_visitors": unique_visitors,
        "inside": inside,
        "outside": max(registered - inside, 0),
        "entries": entries,
        "exits": exits,
        "scan_errors": errors,
    }


@router.get("/statistics/traffic")
async def statistics_traffic(
    event_id: UUID, session: SessionDep, _: Annotated[User, Depends(require_admin)]
) -> dict[str, object]:
    bucket = func.date_trunc("hour", AttendanceEvent.occurred_at)
    rows = (
        await session.execute(
            select(bucket.label("bucket"), AttendanceEvent.event_type, func.count())
            .where(AttendanceEvent.event_id == event_id)
            .group_by(bucket, AttendanceEvent.event_type)
            .order_by(bucket)
        )
    ).all()
    return {"items": [{"bucket": row[0], "event_type": row[1], "count": row[2]} for row in rows]}


@router.get("/statistics/institutes")
async def statistics_institutes(
    event_id: UUID, session: SessionDep, _: Annotated[User, Depends(require_admin)]
) -> dict[str, object]:
    rows = (
        await session.execute(
            select(SourceStudent.institute, func.count())
            .where(SourceStudent.event_id == event_id)
            .group_by(SourceStudent.institute)
            .order_by(func.count().desc())
        )
    ).all()
    return {"items": [{"institute": row[0], "count": row[1]} for row in rows]}


@router.get("/statistics/groups")
async def statistics_groups(
    event_id: UUID, session: SessionDep, _: Annotated[User, Depends(require_admin)]
) -> dict[str, object]:
    rows = (
        await session.execute(
            select(SourceStudent.study_group, func.count())
            .where(SourceStudent.event_id == event_id)
            .group_by(SourceStudent.study_group)
            .order_by(func.count().desc())
        )
    ).all()
    return {"items": [{"group": row[0], "count": row[1]} for row in rows]}


@router.get("/exports/attendance.xlsx")
async def attendance_export(
    request: Request,
    event_id: UUID,
    session: SessionDep,
    user: Annotated[User, Depends(require_admin)],
) -> StreamingResponse:
    event = await session.get(Event, event_id)
    if not event:
        raise AppError("EVENT_NOT_FOUND", "Мероприятие не найдено", 404)
    rows = (
        await session.execute(
            select(AttendanceEvent, Registration, SourceStudent, User)
            .join(Registration, Registration.id == AttendanceEvent.registration_id)
            .join(SourceStudent, SourceStudent.id == Registration.source_student_id)
            .join(User, User.id == AttendanceEvent.operator_id)
            .where(AttendanceEvent.event_id == event_id)
            .order_by(AttendanceEvent.occurred_at)
        )
    ).all()
    workbook = Workbook(write_only=False)
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Посещения"
    headers = [
        "event_name",
        "occurred_at",
        "source_id",
        "name",
        "group",
        "institute",
        "event_type",
        "previous_status",
        "new_status",
        "operator",
        "request_id",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    for attendance, _, student, operator in rows:
        sheet.append(
            [
                _safe_excel(event.name),
                attendance.occurred_at.astimezone(ZoneInfo(event.timezone)).isoformat(),
                _safe_excel(student.source_id),
                _safe_excel(student.full_name),
                _safe_excel(student.study_group),
                _safe_excel(student.institute),
                attendance.event_type.value,
                attendance.previous_status.value,
                attendance.new_status.value,
                _safe_excel(operator.username),
                str(attendance.request_id),
            ]
        )
    for column_index, column in enumerate(sheet.columns, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(len(str(cell.value or "")) for cell in column) + 2, 50
        )
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    add_audit(
        session, request, "ATTENDANCE_EXPORT", "Event", str(event_id), user.id, {"rows": len(rows)}
    )
    await session.commit()
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="attendance.xlsx"'},
    )


@router.post("/events/{event_id}/archive")
async def archive_event(
    event_id: UUID,
    payload: ArchiveEventRequest,
    request: Request,
    session: SessionDep,
    user: Annotated[User, Depends(require_admin)],
) -> dict[str, object]:
    if payload.confirmation_phrase != "АРХИВИРОВАТЬ":
        raise AppError("CONFIRMATION_PHRASE_INVALID", "Введите контрольную фразу АРХИВИРОВАТЬ", 422)
    event = await session.scalar(select(Event).where(Event.id == event_id).with_for_update())
    if not event:
        raise AppError("EVENT_NOT_FOUND", "Мероприятие не найдено", 404)
    event.status, event.registration_enabled, event.scanning_enabled = (
        EventStatus.ARCHIVED,
        False,
        False,
    )
    registration_ids = select(Registration.id).where(Registration.event_id == event_id)
    await session.execute(
        update(QrToken)
        .where(QrToken.registration_id.in_(registration_ids), QrToken.status == QrStatus.ACTIVE)
        .values(status=QrStatus.REVOKED, revoked_at=utc_now())
    )
    add_audit(session, request, "EVENT_ARCHIVE", "Event", str(event.id), user.id)
    await session.commit()
    return {"id": event.id, "status": event.status}
