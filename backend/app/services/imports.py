import hashlib
import io
import zipfile
from dataclasses import dataclass, field
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import select, update
from sqlalchemy.ext.asyncio import AsyncSession

from app.exceptions import AppError
from app.models import ImportBatch, ImportRow, ImportRowError, QrToken, Registration, SourceStudent
from app.models.enums import ErrorSeverity, ImportAction, ImportStatus, QrStatus, ValidationStatus
from app.services.normalization import (
    GROUP_RE,
    has_control_characters,
    normalize_group,
    normalize_institute,
    normalize_name,
)

REQUIRED_HEADERS = ("id", "name", "group", "institute")


@dataclass
class RowIssue:
    code: str
    message: str
    severity: ErrorSeverity


@dataclass
class ParsedRow:
    row_number: int
    source_id: str | None
    name: str | None
    group: str | None
    institute: str | None
    normalized_name: str | None = None
    normalized_group: str | None = None
    normalized_institute: str | None = None
    issues: list[RowIssue] = field(default_factory=list)
    action: ImportAction = ImportAction.ERROR

    @property
    def has_errors(self) -> bool:
        return any(issue.severity == ErrorSeverity.ERROR for issue in self.issues)


def _source_id(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            number = Decimal(str(value))
        except InvalidOperation:
            return None
        if number != number.to_integral_value() or number <= 0:
            return None
        return str(number.quantize(Decimal("1")))
    result = str(value).strip()
    return result if result else None


def _cell_text(value: Any) -> str | None:
    if value is None:
        return None
    result = str(value).strip()
    return result if result else None


def validate_xlsx_container(data: bytes, max_rows: int) -> None:
    if len(data) < 4 or data[:4] != b"PK\x03\x04":
        raise AppError("IMPORT_INVALID_XLSX", "Файл не является корректным XLSX", 422)
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            entries = archive.infolist()
            if not entries or len(entries) > 2_000:
                raise AppError("IMPORT_ZIP_BOMB", "Структура XLSX небезопасна", 422)
            total_uncompressed = sum(entry.file_size for entry in entries)
            total_compressed = max(sum(entry.compress_size for entry in entries), 1)
            if (
                total_uncompressed > 200 * 1024 * 1024
                or total_uncompressed / total_compressed > 100
            ):
                raise AppError(
                    "IMPORT_ZIP_BOMB", "XLSX превышает безопасный коэффициент сжатия", 422
                )
            names = {entry.filename for entry in entries}
            if "[Content_Types].xml" not in names or not any(
                name.startswith("xl/worksheets/") for name in names
            ):
                raise AppError("IMPORT_INVALID_XLSX", "Структура XLSX повреждена", 422)
            if any("vbaproject" in name.casefold() for name in names):
                raise AppError(
                    "IMPORT_MACROS_NOT_ALLOWED", "Файлы с макросами не поддерживаются", 422
                )
    except zipfile.BadZipFile as exc:
        raise AppError("IMPORT_INVALID_XLSX", "Архив XLSX повреждён", 422) from exc


def parse_workbook(data: bytes, max_rows: int) -> tuple[list[ParsedRow], list[RowIssue]]:
    validate_xlsx_container(data, max_rows)
    try:
        workbook = load_workbook(
            io.BytesIO(data), read_only=True, data_only=False, keep_links=False
        )
    except (InvalidFileException, OSError, ValueError, KeyError) as exc:
        raise AppError("IMPORT_INVALID_XLSX", "Не удалось безопасно прочитать XLSX", 422) from exc
    if not workbook.worksheets:
        raise AppError("IMPORT_FILE_EMPTY", "В файле нет листов", 422)
    sheet = workbook.worksheets[0]
    iterator = sheet.iter_rows()
    try:
        header_cells = next(iterator)
    except StopIteration as exc:
        raise AppError("IMPORT_FILE_EMPTY", "Файл пуст", 422) from exc
    headers = [
        str(cell.value).strip().casefold() if cell.value is not None else ""
        for cell in header_cells
    ]
    missing = [name for name in REQUIRED_HEADERS if name not in headers]
    if missing:
        raise AppError(
            "IMPORT_REQUIRED_COLUMN_MISSING",
            f"Отсутствуют обязательные колонки: {', '.join(missing)}",
            422,
            {"missing": missing},
        )
    header_map = {name: headers.index(name) for name in REQUIRED_HEADERS}
    batch_issues = (
        [
            RowIssue(
                "IMPORT_EXTRA_COLUMNS",
                "Дополнительные колонки проигнорированы",
                ErrorSeverity.WARNING,
            )
        ]
        if any(name and name not in REQUIRED_HEADERS for name in headers)
        else []
    )
    parsed: list[ParsedRow] = []
    for row_number, cells in enumerate(iterator, start=2):
        if row_number - 1 > max_rows:
            raise AppError("IMPORT_TOO_MANY_ROWS", f"Допустимо не более {max_rows} строк", 413)
        chosen = {
            name: cells[index] if index < len(cells) else None for name, index in header_map.items()
        }
        if all(cell is None or cell.value is None for cell in chosen.values()):
            continue
        formula_columns = [
            name for name, cell in chosen.items() if cell is not None and cell.data_type == "f"
        ]
        row = ParsedRow(
            row_number=row_number,
            source_id=_source_id(chosen["id"].value if chosen["id"] else None),
            name=_cell_text(chosen["name"].value if chosen["name"] else None),
            group=_cell_text(chosen["group"].value if chosen["group"] else None),
            institute=_cell_text(chosen["institute"].value if chosen["institute"] else None),
        )
        if formula_columns:
            row.issues.append(
                RowIssue(
                    "IMPORT_FORMULA_NOT_ALLOWED",
                    f"Формулы запрещены: {', '.join(formula_columns)}",
                    ErrorSeverity.ERROR,
                )
            )
        for value, code, label in (
            (row.source_id, "IMPORT_SOURCE_ID_MISSING", "id"),
            (row.name, "IMPORT_NAME_MISSING", "ФИО"),
            (row.group, "IMPORT_GROUP_MISSING", "группа"),
            (row.institute, "IMPORT_INSTITUTE_MISSING", "институт"),
        ):
            if not value:
                row.issues.append(
                    RowIssue(code, f"Не заполнено обязательное поле: {label}", ErrorSeverity.ERROR)
                )
            elif has_control_characters(value):
                row.issues.append(
                    RowIssue(
                        "IMPORT_CONTROL_CHARACTER",
                        f"Поле {label} содержит управляющий символ",
                        ErrorSeverity.ERROR,
                    )
                )
        if row.source_id and len(row.source_id) > 128:
            row.issues.append(
                RowIssue("IMPORT_SOURCE_ID_INVALID", "id длиннее 128 символов", ErrorSeverity.ERROR)
            )
        if row.name:
            row.normalized_name = normalize_name(row.name)
        if row.group:
            row.normalized_group = normalize_group(row.group)
            if not GROUP_RE.fullmatch(row.normalized_group):
                row.issues.append(
                    RowIssue(
                        "IMPORT_GROUP_FORMAT_SUSPICIOUS",
                        "Необычный формат учебной группы",
                        ErrorSeverity.WARNING,
                    )
                )
        if row.institute:
            row.normalized_institute = normalize_institute(row.institute)
        parsed.append(row)
    if not parsed:
        raise AppError("IMPORT_FILE_EMPTY", "Файл не содержит строк студентов", 422)
    by_source: dict[str, list[ParsedRow]] = {}
    for row in parsed:
        if row.source_id:
            by_source.setdefault(row.source_id, []).append(row)
    for source_id, rows in by_source.items():
        if len(rows) > 1:
            numbers = [row.row_number for row in rows]
            for row in rows:
                row.issues.append(
                    RowIssue(
                        "IMPORT_SOURCE_ID_DUPLICATE",
                        f"Повтор id={source_id} в строках {numbers}",
                        ErrorSeverity.ERROR,
                    )
                )
    by_person: dict[tuple[str | None, str | None, str | None], list[ParsedRow]] = {}
    for row in parsed:
        if not row.has_errors:
            by_person.setdefault(
                (row.normalized_name, row.normalized_group, row.normalized_institute), []
            ).append(row)
    for rows in by_person.values():
        if len({row.source_id for row in rows}) > 1:
            for row in rows:
                row.issues.append(
                    RowIssue(
                        "IMPORT_DUPLICATE_PERSON",
                        "Разные id имеют одинаковые ФИО, группу и институт",
                        ErrorSeverity.WARNING,
                    )
                )
    return parsed, batch_issues


async def stage_import(
    session: AsyncSession, batch: ImportBatch, parsed: list[ParsedRow], batch_issues: list[RowIssue]
) -> None:
    existing = {
        student.source_id: student
        for student in (
            await session.scalars(
                select(SourceStudent).where(SourceStudent.event_id == batch.event_id)
            )
        ).all()
    }
    incoming: set[str] = set()
    counts = {action: 0 for action in ImportAction}
    errors = warnings = valid = 0
    for parsed_row in parsed:
        if parsed_row.source_id:
            incoming.add(parsed_row.source_id)
        current = existing.get(parsed_row.source_id or "")
        if parsed_row.has_errors:
            parsed_row.action = ImportAction.ERROR
        elif not current:
            parsed_row.action = ImportAction.CREATE
        elif (current.full_name, current.study_group, current.institute) != (
            parsed_row.name,
            parsed_row.group,
            parsed_row.institute,
        ):
            parsed_row.action = ImportAction.UPDATE
            parsed_row.issues.append(
                RowIssue(
                    "IMPORT_EXISTING_STUDENT_CHANGED",
                    "Данные существующего студента будут обновлены",
                    ErrorSeverity.WARNING,
                )
            )
        else:
            parsed_row.action = ImportAction.UNCHANGED
        status = (
            ValidationStatus.ERROR
            if parsed_row.has_errors
            else ValidationStatus.WARNING
            if parsed_row.issues
            else ValidationStatus.VALID
        )
        model = ImportRow(
            import_batch_id=batch.id,
            row_number=parsed_row.row_number,
            source_id=parsed_row.source_id,
            name=parsed_row.name,
            group=parsed_row.group,
            institute=parsed_row.institute,
            normalized_name=parsed_row.normalized_name,
            normalized_group=parsed_row.normalized_group,
            normalized_institute=parsed_row.normalized_institute,
            action=parsed_row.action,
            validation_status=status,
            created_at=datetime.now(UTC),
        )
        session.add(model)
        await session.flush()
        for issue in parsed_row.issues:
            session.add(
                ImportRowError(
                    import_batch_id=batch.id,
                    import_row_id=model.id,
                    row_number=parsed_row.row_number,
                    error_code=issue.code,
                    error_message=issue.message,
                    severity=issue.severity,
                    created_at=datetime.now(UTC),
                )
            )
        counts[parsed_row.action] += 1
        errors += int(parsed_row.has_errors)
        warnings += int(any(issue.severity == ErrorSeverity.WARNING for issue in parsed_row.issues))
        valid += int(not parsed_row.has_errors)
    missing = [
        student
        for source_id, student in existing.items()
        if student.is_active and source_id not in incoming
    ]
    for index, student in enumerate(missing, start=1):
        session.add(
            ImportRow(
                import_batch_id=batch.id,
                row_number=-index,
                source_id=student.source_id,
                name=student.full_name,
                group=student.study_group,
                institute=student.institute,
                normalized_name=student.normalized_full_name,
                normalized_group=student.normalized_study_group,
                normalized_institute=student.normalized_institute,
                action=ImportAction.DEACTIVATE,
                validation_status=ValidationStatus.WARNING,
                created_at=datetime.now(UTC),
            )
        )
    batch.total_rows = len(parsed)
    batch.valid_rows, batch.error_rows, batch.warning_rows = (
        valid,
        errors,
        warnings + len(batch_issues) + len(missing),
    )
    batch.created_rows, batch.updated_rows, batch.unchanged_rows, batch.deactivated_rows = (
        counts[ImportAction.CREATE],
        counts[ImportAction.UPDATE],
        counts[ImportAction.UNCHANGED],
        len(missing),
    )
    batch.status = ImportStatus.READY_TO_CONFIRM if errors == 0 else ImportStatus.VALIDATED
    batch.validated_at = datetime.now(UTC)
    updated_marker = batch.updated_at.isoformat() if batch.updated_at else ""
    preview_source = f"{batch.file_hash}:{updated_marker}:{len(existing)}"
    batch.preview_version = hashlib.sha256(preview_source.encode()).hexdigest()


async def apply_import(
    session: AsyncSession, batch: ImportBatch, confirm_deactivations: bool
) -> None:
    rows = (
        await session.scalars(
            select(ImportRow)
            .where(ImportRow.import_batch_id == batch.id)
            .order_by(ImportRow.row_number)
        )
    ).all()
    now = datetime.now(UTC)
    for row in rows:
        if row.action in (ImportAction.CREATE, ImportAction.UPDATE):
            if not all(
                (
                    row.source_id,
                    row.name,
                    row.group,
                    row.institute,
                    row.normalized_name,
                    row.normalized_group,
                    row.normalized_institute,
                )
            ):
                raise AppError("IMPORT_STAGING_INVALID", "Промежуточные данные повреждены", 409)
            assert row.source_id is not None
            assert row.name is not None
            assert row.group is not None
            assert row.institute is not None
            assert row.normalized_name is not None
            assert row.normalized_group is not None
            assert row.normalized_institute is not None
        if row.action == ImportAction.CREATE:
            session.add(
                SourceStudent(
                    event_id=batch.event_id,
                    source_id=row.source_id,
                    full_name=row.name,
                    normalized_full_name=row.normalized_name,
                    study_group=row.group,
                    normalized_study_group=row.normalized_group,
                    institute=row.institute,
                    normalized_institute=row.normalized_institute,
                    source_row_number=row.row_number,
                    import_batch_id=batch.id,
                    is_active=True,
                )
            )
        elif row.action == ImportAction.UPDATE:
            assert row.name is not None
            assert row.group is not None
            assert row.institute is not None
            assert row.normalized_name is not None
            assert row.normalized_group is not None
            assert row.normalized_institute is not None
            student = await session.scalar(
                select(SourceStudent)
                .where(
                    SourceStudent.event_id == batch.event_id,
                    SourceStudent.source_id == row.source_id,
                )
                .with_for_update()
            )
            if student:
                student.full_name, student.normalized_full_name = row.name, row.normalized_name
                student.study_group, student.normalized_study_group = (
                    row.group,
                    row.normalized_group,
                )
                student.institute, student.normalized_institute = (
                    row.institute,
                    row.normalized_institute,
                )
                (
                    student.source_row_number,
                    student.import_batch_id,
                    student.is_active,
                    student.deactivated_at,
                ) = row.row_number, batch.id, True, None
        elif row.action == ImportAction.DEACTIVATE and confirm_deactivations:
            student = await session.scalar(
                select(SourceStudent)
                .where(
                    SourceStudent.event_id == batch.event_id,
                    SourceStudent.source_id == row.source_id,
                )
                .with_for_update()
            )
            if student:
                student.is_active, student.deactivated_at = False, now
                registration_ids = select(Registration.id).where(
                    Registration.source_student_id == student.id
                )
                await session.execute(
                    update(QrToken)
                    .where(
                        QrToken.registration_id.in_(registration_ids),
                        QrToken.status == QrStatus.ACTIVE,
                    )
                    .values(status=QrStatus.REVOKED, revoked_at=now)
                )
